from __future__ import annotations
import io
import logging
import re

logger = logging.getLogger(__name__)

_DEFAULT_COL_WIDTH = 8.0
_DEFAULT_ROW_HEIGHT = 15.0


# ── Helpers ───────────────────────────────────────────────────────────────────

def _col_char_widths(ws) -> list[float]:
    from openpyxl.utils import get_column_letter
    widths = []
    for i in range(1, ws.max_column + 1):
        letter = get_column_letter(i)
        dim = ws.column_dimensions.get(letter)
        widths.append(dim.width if dim and dim.width else _DEFAULT_COL_WIDTH)
    return widths


def _row_pt_heights(ws) -> list[float]:
    heights = []
    for i in range(1, ws.max_row + 1):
        dim = ws.row_dimensions.get(i)
        heights.append(dim.height if dim and dim.height else _DEFAULT_ROW_HEIGHT)
    return heights


def _cumulative(sizes: list[float]) -> list[float]:
    acc = [0.0]
    for s in sizes:
        acc.append(acc[-1] + s)
    return acc


def _find_value_cells(ws) -> dict[int, tuple[int, int]]:
    """Return {field_number: (row, col)} for each NTS field's value cell top-left."""
    panel_boundary = ws.max_column // 2

    field_num_cells: list[tuple[int, int, int]] = []
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is None:
                continue
            val = str(cell.value).strip()
            if re.match(r'^\d{1,3}$', val):
                num = int(val)
                if 1 <= num <= 200:
                    field_num_cells.append((cell.row, cell.column, num))

    result: dict[int, tuple[int, int]] = {}
    for (row, num_col, num) in field_num_cells:
        search_max = panel_boundary if num_col <= panel_boundary else ws.max_column
        best_range = None
        best_col = -1
        for mr in ws.merged_cells.ranges:
            if mr.min_row <= row <= mr.max_row:
                if num_col < mr.min_col <= search_max:
                    if mr.min_col > best_col:
                        best_col = mr.min_col
                        best_range = mr
        if best_range:
            result[num] = (best_range.min_row, best_range.min_col)

    return result


def _eval_formula(formula: str, values: dict[str, float | None]) -> float | None:
    """Evaluate a formula string like '01+02-03' using current field values."""
    if not formula:
        return None
    expr = str(formula)
    for fid in sorted(values.keys(), key=len, reverse=True):
        val = values.get(fid)
        expr = re.sub(r'\b' + re.escape(fid) + r'\b', str(val if val is not None else 0), expr)
    # Safety check: only digits and arithmetic operators remain
    if re.search(r'[^0-9+\-*/().\s]', expr):
        return None
    try:
        result = eval(expr)  # noqa: S307
        return float(result) if isinstance(result, (int, float)) else None
    except Exception:
        return None


def _compute_all_values(fields: list[dict], input_values: dict) -> dict[str, float | None]:
    """Compute all field values — mirrors useFormCalculation hook (two passes)."""
    values: dict[str, float | None] = {f['id']: input_values.get(f['id']) for f in fields}
    for _ in range(2):
        for field in fields:
            if field.get('type') != 'calculated' or not field.get('formula'):
                continue
            values[field['id']] = _eval_formula(field['formula'], values)
    return values


# ── Public API ────────────────────────────────────────────────────────────────

def extract_bbox_from_excel(
    excel_bytes: bytes,
    fields: list[dict],
    sheet_index: int = 0,
    pdf_bytes: bytes | None = None,
) -> list[dict]:
    """Parse Excel and add normalized bbox to each field (kept for legacy use)."""
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
        sheets = wb.worksheets
        if not sheets:
            return fields
        ws = sheets[sheet_index] if sheet_index < len(sheets) else sheets[0]

        page_w_pt, page_h_pt = 595.0, 842.0
        if pdf_bytes:
            try:
                import fitz
                doc = fitz.open(stream=pdf_bytes, filetype="pdf")
                if doc:
                    r = doc[0].rect
                    page_w_pt, page_h_pt = r.width, r.height
                    doc.close()
            except Exception:
                pass

        col_chars = _col_char_widths(ws)
        total_chars = sum(col_chars)
        char_to_pt = page_w_pt / total_chars if total_chars else 1.0
        col_pos = _cumulative([w * char_to_pt for w in col_chars])

        row_pts = _row_pt_heights(ws)
        total_row_pts = sum(row_pts)
        row_scale = page_h_pt / total_row_pts if total_row_pts else 1.0
        row_pos = _cumulative([h * row_scale for h in row_pts])

        value_cells = _find_value_cells(ws)

        def merged_bbox(mr):
            x = col_pos[mr.min_col - 1]
            y = row_pos[mr.min_row - 1]
            return {
                "page": 0,
                "x": x / page_w_pt,
                "y": y / page_h_pt,
                "w": (col_pos[mr.max_col] - x) / page_w_pt,
                "h": (row_pos[mr.max_row] - y) / page_h_pt,
            }

        # Build num→merged_range map
        panel_boundary = ws.max_column // 2
        field_num_to_bbox: dict[int, dict] = {}
        for row in ws.iter_rows():
            for cell in row:
                if cell.value is None:
                    continue
                val = str(cell.value).strip()
                if re.match(r'^\d{1,3}$', val):
                    num = int(val)
                    if 1 <= num <= 200:
                        num_col = cell.column
                        search_max = panel_boundary if num_col <= panel_boundary else ws.max_column
                        best_range = None
                        best_col = -1
                        for mr in ws.merged_cells.ranges:
                            if mr.min_row <= cell.row <= mr.max_row:
                                if num_col < mr.min_col <= search_max:
                                    if mr.min_col > best_col:
                                        best_col = mr.min_col
                                        best_range = mr
                        if best_range:
                            field_num_to_bbox[num] = merged_bbox(best_range)

        updated = []
        for field in fields:
            field = dict(field)
            try:
                num = int(str(field.get("id", "")).strip())
                if num in field_num_to_bbox:
                    field["bbox"] = field_num_to_bbox[num]
            except (ValueError, TypeError):
                pass
            updated.append(field)
        return updated

    except Exception as e:
        logger.error("Excel bbox extraction failed: %s", e)
        return fields


def fill_excel_with_values(
    excel_bytes: bytes,
    fields: list[dict],
    input_values: dict,
    sheet_index: int = 0,
) -> bytes:
    """Fill field values into the Excel value cells and return modified bytes."""
    from openpyxl import load_workbook
    wb = load_workbook(io.BytesIO(excel_bytes))
    sheets = wb.worksheets
    ws = sheets[sheet_index] if sheet_index < len(sheets) else sheets[0]

    computed = _compute_all_values(fields, input_values)
    value_cells = _find_value_cells(ws)

    for field in fields:
        fid = field.get("id", "")
        try:
            num = int(str(fid).strip())
        except (ValueError, TypeError):
            continue
        if num not in value_cells:
            continue
        val = computed.get(fid)
        if val is None:
            continue
        row, col = value_cells[num]
        cell = ws.cell(row=row, column=col)
        # Format as integer if whole number, else 2 decimal places
        cell.value = int(val) if val == int(val) else round(val, 2)
        # Right-align the value
        from openpyxl.styles import Alignment
        cell.alignment = Alignment(horizontal="right", vertical="center")

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def excel_to_html(excel_bytes: bytes, sheet_index: int = 0) -> str:
    """Convert Excel to HTML string using xlsx2html."""
    import xlsx2html
    buf_in = io.BytesIO(excel_bytes)
    buf_out = io.StringIO()
    xlsx2html.xlsx2html(buf_in, buf_out, locale="ko_KR", sheet=sheet_index)
    html = buf_out.getvalue()
    # Inject base styles for readability
    style = """
    <style>
      body { margin: 0; padding: 16px; background: #f1f5f9; font-family: 'Malgun Gothic', sans-serif; }
      table { border-collapse: collapse; background: #fff; }
      td, th { font-family: 'Malgun Gothic', 'Apple SD Gothic Neo', sans-serif !important; }
    </style>
    """
    return html.replace("</head>", style + "</head>", 1) if "</head>" in html else style + html
